import json
import os
import openpyxl


class Utils:
    
    @staticmethod
    def read_urls_from_txt(file_path):
        """Reads URLs from a text file, removes newline characters, and returns them as a list."""
        try:
            with open(file_path, "r") as file:
                urls = [line.strip("\n") for line in file.readlines()]  # Remove "\n" from each line
            return urls
        except FileNotFoundError:
            print(f"Error: File not found: {file_path}")
            return []
        except Exception as e:
            print(f"An error occurred while reading the file: {e}")
            return []

    @staticmethod
    def filter_emails_from_json(file_path):
        """
        Reads all the emails from the json file and deletes mutltiple
        and empty arrays making a text file for the emails of the json file
        """
        try:
            with open(file_path, "r") as file:
                data = json.load(file)
                emails = []
                for emails_dict in data:
                    emails.extend(emails_dict["emails"])
                    
                emails = list(set(emails))  # Remove duplicates
                
                return emails
            
        except FileNotFoundError:
            print(f"Error: File not found: {file_path}")
            return []
        except Exception as e:
            print(f"An error occurred while reading the file: {e}")
            return []
                    
    @staticmethod
    def write_emails_to_txt(file_path, emails):
        """Writes the emails to a text file, separated by newlines."""
        try:
            with open(file_path, "w") as file:
                file.writelines(email + "\n" for email in emails)
        except Exception as e:
            print(f"An error occurred while writing the file: {e}")
            return False
        return True
    
    @staticmethod
    def create_dir_if_not_exists(directory):
        """
        Creates a directory if it doesn't already exist.

        Args:
            directory (str): The path to the directory to create.
        """
        if not os.path.exists(directory):
            try:
                os.makedirs(directory)
                print(f"Created directory: {directory}")
            except OSError as e:
                print(f"Error creating directory: {e}")
        
    def read_urls_from_excel(filename):
        """
        Reads university departments and their links from the specified Excel sheet
        and creates a dictionary with the structure:

        {link: cell_location}

        Args:
            filename (str): Path to the Excel file containing the data.

        Returns:
            dict: Dictionary mapping department links to cell locations.
        """

        link_cell_dict = {}
        wb = openpyxl.load_workbook(filename)
        sheet = wb.active  # Assuming the data is in the first sheet

        # Start from the second row (skipping the header)
        for row_num in range(2, sheet.max_row + 1):
            # Check if cell value is not None before splitting
            university_info = sheet.cell(row=row_num, column=1).value
            if university_info:
                # Extract university name (assuming you don't need number)
                university_name = university_info.strip()

                # Move to the first department link cell (column 2)
                department_cell = sheet.cell(row=row_num, column=2)

                # Loop until the end of the row (using max column)
                for col_num in range(2, sheet.max_column + 1):
                    department_cell = sheet.cell(row=row_num, column=col_num)
                    if department_cell.value:
                        link = department_cell.value
                        cell_location = (department_cell.row, department_cell.column)
                        link_cell_dict[link] = cell_location

            # Continue to the next row regardless of empty departments
            continue

        return link_cell_dict
    
    
        
    def insert_data_to_sheet(filename, row, col, data):
        """
        Inserts data into the specified cell of an Excel sheet.

        Args:
            filename (str): Path to the Excel file.
            row (int): Row number (starting from 1).
            col (int): Column number (starting from 1).
            data (any): Data to be inserted into the cell.
        """

        wb = openpyxl.load_workbook(filename)
        sheet = wb.active  # Assuming data goes into the first sheet

        sheet.cell(row=row, column=col).value = data

        # Save the changes to the Excel file
        wb.save(filename)
             
    @staticmethod       
    def remove_white_spaces_in_emails(file_path):
        """
        Removes any whitespace in between each email in the specified file.

        Args:
            file_path (str): Path to the file containing the emails.
        """

        with open(file_path, "r") as file:
            lines = file.readlines()

        with open(file_path, "w") as file:
            for line in lines:
                email = line.strip()
                for char in email:
                    if char == " ":
                        email = email.replace(" ", "")
                file.write(email + "\n")
                
        with open(file_path, "r") as file:
            return file.read()

    @staticmethod
    def main():
        excel_mapping = Utils.read_urls_from_excel("emails.xlsx")
    
        # Getting key value pairs in dictionary
        for url, cell_location in excel_mapping.items():
            # print(url, cell_location)
            
            file = f"emails/{cell_location[0]}{cell_location[1]}.txt"
            Utils.insert_data_to_sheet("emails.xlsx", cell_location[0]+1, cell_location[1],Utils.remove_white_spaces_in_emails(file))
            # Utils.remove_white_spaces_in_emails(f"emails/{cell_location[0]}{cell_location[1]}.txt")
            
        # Utils.write_emails_to_txt("emails/24.txt",Utils.filter_emails_from_json("feed/24.json"))        
        
        
if __name__ == "__main__":
    # print(Utils.remove_white_spaces_in_emails(f"emails/122.txt"))
    Utils.main()
    # print(Utils.read_urls_from_excel("emails.xlsx"))